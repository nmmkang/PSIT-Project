"""Take data from main GUI and database file, check data for errors/duplicates, write data to spreadsheet."""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from os.path import isfile
from datetime import datetime

#path of database file
data_path = 'data' + '.xlsx'

#if data file does not exist, create a new one.
if not isfile('./' + data_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Passengers'

    ws.append(['Name', 'Name', 'Age', 'Tel', 'Email', 'Start', 'Time', 'Date',
            'Destination', 'Time', 'Date', 'Class', 'Seat', 'Timestamp'])

    black = '00000000'
    thin = Side(border_style='thin', color=black)

    for i in range(1, 15):
        ws.cell(row=1, column=i).font = Font(bold=True)
        ws.cell(row=1, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    green1 = '00C6E0B4'
    green2 = '00E2EFDA'
    blue = '00B4C6E7'
    blue2 = '0092CDDC'
    red1 = '00F4B084'
    red2 = '00F8CBAD'
    yellow1 = '00FFD966'
    yellow2 = '00FFE699'
    lime = '00CCFF33'

    ws['A1'].fill = PatternFill(patternType='solid', fgColor=green2)
    ws['B1'].fill = PatternFill(patternType='solid', fgColor=green1)
    ws['C1'].fill = PatternFill(patternType='solid', fgColor=green2)
    ws['D1'].fill = PatternFill(patternType='solid', fgColor=blue)
    ws['E1'].fill = PatternFill(patternType='solid', fgColor=blue)
    ws['F1'].fill = PatternFill(patternType='solid', fgColor=red1)
    ws['G1'].fill = PatternFill(patternType='solid', fgColor=red2)
    ws['H1'].fill = PatternFill(patternType='solid', fgColor=red2)
    ws['I1'].fill = PatternFill(patternType='solid', fgColor=yellow1)
    ws['J1'].fill = PatternFill(patternType='solid', fgColor=yellow2)
    ws['K1'].fill = PatternFill(patternType='solid', fgColor=yellow2)
    ws['L1'].fill = PatternFill(patternType='solid', fgColor=lime)
    ws['M1'].fill = PatternFill(patternType='solid', fgColor=lime)
    ws['N1'].fill = PatternFill(patternType='solid', fgColor=blue2)

    ws.column_dimensions['A'].width = 64/6
    ws.column_dimensions['B'].width = 216/6
    ws.column_dimensions['C'].width = 64/6
    ws.column_dimensions['D'].width = 115/6
    ws.column_dimensions['E'].width = 115/6
    ws.column_dimensions['F'].width = 150/6
    ws.column_dimensions['G'].width = 85/6
    ws.column_dimensions['H'].width = 85/6
    ws.column_dimensions['I'].width = 150/6
    ws.column_dimensions['J'].width = 85/6
    ws.column_dimensions['K'].width = 85/6
    ws.column_dimensions['L'].width = 100/6
    ws.column_dimensions['M'].width = 64/6
    ws.column_dimensions['N'].width = 144/6

    wb.save(data_path)

#load workbook
wb = load_workbook(data_path)
ws = wb.active

#อ่านข้อมูลใน sheet มาเก็บไว้ใน list เพื่อใช้เปรีบยเทียบดูข้อมูลซ้ำ
#sheet มี 13 columes
#1:name_start, 2:name,           3:age,        4:tel,  5:email,
#6:start,      7:start_time,     8:start_date, 9:dest, 10:dest_time,
#11:dest_date, 12:class_airport, 13:seat
current_data = [
    list(list(ws.iter_cols(min_row=2, min_col=col, max_col=col, values_only=True))[0])
    for col in range(1, 14)
]

#print all datas, probably use for debug
for col in current_data:
    print(col)

def timestamp():
    """Get current time. %Y/%m/%d %H:%M:%S"""
    return datetime.now().strftime('%Y/%m/%d %H:%M:%S')

def write(data):
    """Write data to spreadsheet."""
    data['timestamp'] = timestamp()
    ws.append(list(data.values()))
    wb.save(data_path)

#* USER INPUT ERROR
#? Check if: - input field is blank
#?           - input is in the wrong format


def check_error(name_start, name, age, tel, email, origin, time_start,
                date_departure, destination, time_return, date_return,
                class_seat, seat_chr, seat_num):
    """Check if any input field is error."""
    return sum([error_blank(name_start), error_blank(name), bool(error_numeric(age)),
                bool(error_numeric(tel)), bool(error_email(email)), error_blank(origin),
                error_blank(time_start), bool(error_date(date_departure)),
                error_blank(destination), error_blank(time_return),
                bool(error_date(date_return)), error_blank(class_seat),
                error_blank(seat_chr), error_blank(seat_num)]) != 0

def error_blank(var):
    """Check most entries/comboboxes/radiobuttons if it blanks."""
    return var == '' or var.isspace()

def error_numeric(num):
    """Check entry if it's blank/not a number."""
    if num == '':
        return 'blank'
    elif not num.isnumeric():
        return 'num'
    return False

def error_email(email):
    """Check email if it's blank/wrong format."""
    if email == '':
        return 'blank'
    elif email.count('@') != 1 or email.count('.') < 1:
        return 'format'
    return False

def error_date(date):
    """Check dates if blank/wrong format."""
    if date == '':
        return 'blank'
    if date[2:6:3] != '--':
        return 'format'
    if not 12 >= int(date[3:5]) >= 1:
        return 'month'
    if int(date[:2]) < 1:
        return 'day'
    elif int(date[3:5]) in [1, 3, 5, 7, 8, 10, 12] and int(date[:2]) > 31: #คม
        return 'day'
    elif int(date[3:5]) in [4, 6, 9, 11] and int(date[:2]) > 30: #ยน
        return 'day'
    return False

#* DUPLICATE CHECKING
#? Check if any input is duplicated with one recorded in the database.

def check_duplicate(name, tel, email, origin, date_departure, time_start, destination,
                    date_return, time_return, seat):
    """Check if any input is duplicated with database."""
    return sum([dupl_name(name), dupl_tel(tel), dupl_email(email),
                dupl_seat(seat, origin, date_departure, time_start, 'start'),
                dupl_seat(seat, destination, date_return, time_return, 'end')]) != 0

#0:name_start, 1:name,           2:age,        3:tel,  4:email,
#5:start,      6:start_time,     7:start_date, 8:dest, 9:dest_time,
#10:dest_date, 11:class_airport, 12:seat

def dupl_name(name):
    """Check for name duplicates."""
    return name in current_data[1]

def dupl_tel(tel):
    """Check for tel duplicates."""
    return tel in current_data[3]

def dupl_email(email):
    """Check for email duplicates."""
    return email in current_data[4]

def dupl_seat(seat, place, date, time, check):
    #print(current_data[12])
    if check == 'start':
        col = 5
    elif check == 'end':
        col = 8
    used_seats_start = [ #if start_date in sheet == one in data, add seat into this list
        current_data[12][i]
        for i in range(len(current_data[12]))
        if current_data[col][i] == place and
        current_data[col+1][i] == time and
        current_data[col+2][i] == date
    ]
    print('[CONSOLE] Seats used, %s, %s, %s, %s:' %(check, date, place, time))
    print(used_seats_start)
    return seat in used_seats_start

